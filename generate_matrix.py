import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2026版支撑关系矩阵"

requirements = {
    "1、工程知识": [("1.1","数学基础"), ("1.2","自然科学基础"), ("1.3","工程基础"), ("1.4","专业知识")],
    "2、问题分析": [("2.1","识别与判断"), ("2.2","建模与表达"), ("2.3","分析与求解"), ("2.4","文献调研与综述")],
    "3、设计/开发": [("3.1","需求分析"), ("3.2","方案设计"), ("3.3","系统实现"), ("3.4","创新与优化")],
    "4、研究":      [("4.1","实验设计与调查"), ("4.2","数据采集与处理"), ("4.3","结果分析与解释"), ("4.4","结论与应用")],
    "5、使用工具":  [("5.1","软件工具"), ("5.2","硬件/仪器"), ("5.3","仿真与模拟")],
    "6、工程与社会":[("6.1","工程与社会关系"), ("6.2","社会责任与法规")],
    "7、环境与可持续发展":[("7.1","环境意识"), ("7.2","可持续发展")],
    "8、职业规范":  [("8.1","职业道德"), ("8.2","法律法规"), ("8.3","学术诚信")],
    "9、个人与团队":[("9.1","个人角色"), ("9.2","团队协作"), ("9.3","组织协调")],
    "10、沟通":     [("10.1","书面表达"), ("10.2","口头表达"), ("10.3","国际视野与跨文化")],
    "11、项目管理": [("11.1","项目管理知识"), ("11.2","经济决策"), ("11.3","组织实施")],
    "12、终身学习": [("12.1","自主学习"), ("12.2","适应发展")]
}

all_indicators = []
for req_name, indicators in requirements.items():
    for code, desc in indicators:
        all_indicators.append(code)

courses = []

def add(board, name, credit, nature, mapping):
    courses.append((board, name, credit, nature, mapping))

B = "人文社会科学类通识教育课程"

add(B, "大学英语(普通类)1-4", "8", "必修", {"10.1":"H","10.2":"M","10.3":"H","12.1":"M"})
add(B, "大学体育1-6", "4", "必修", {"9.2":"M"})
add(B, "形势与政策1-8", "2", "必修", {"6.1":"L","7.1":"L","8.2":"M"})
add(B, "思想道德与法治", "3", "必修", {"6.2":"H","7.1":"M","8.1":"H","8.2":"H","8.3":"H"})
add(B, "中国近现代史纲要", "3", "必修", {"6.1":"M","8.1":"M"})
add(B, "马克思主义基本原理", "3", "必修", {"7.2":"M","8.1":"M","12.2":"L"})
add(B, "毛泽东思想和中国特色社会主义理论体系概论", "3", "必修", {"6.1":"M","6.2":"M","8.1":"H"})
add(B, "习近平新时代中国特色社会主义思想概论", "3", "必修", {"6.1":"M","6.2":"H","7.2":"M","8.1":"H"})
add(B, "国防教育", "2", "必修", {"6.2":"M","9.2":"M"})
add(B, "劳动教育", "1", "必修", {"7.1":"M","9.1":"M","11.3":"L"})
add(B, "大学生心理健康教育", "2", "必修", {"8.1":"M","9.1":"M","12.1":"L"})
add(B, "大学生职业发展与就业指导(1)(2)", "1", "必修", {"8.1":"M","9.1":"M","12.1":"M","12.2":"H"})
add(B, "沟通与写作", "2", "必修", {"9.3":"H","10.1":"H","10.2":"H"})
add(B, "创新创业基础", "1", "必修", {"3.4":"M","9.3":"M","11.1":"H","11.2":"M","12.2":"M"})
add(B, "国家安全教育", "1", "必修", {"6.2":"H","8.2":"M"})
add(B, "信息素养与终身学习", "0.5", "必修", {"2.4":"M","5.1":"H","12.1":"H","12.2":"M"})
add(B, "入学教育与新生研讨", "0", "必修", {"12.1":"M"})
add(B, "创新创业实践(通识限选)", "2", "限选", {"3.4":"M","9.3":"M","11.1":"M","11.2":"L","12.2":"H"})
add(B, "文化史(通识限选)", "2", "限选", {"6.1":"M","10.3":"M"})
add(B, "美育导论(通识限选)", "2", "限选", {"7.1":"M"})
add(B, "通识教育选修(其他任选)", "4", "任选", {"6.1":"L","7.1":"L","10.3":"M","12.1":"L"})

B = "数学与自然科学类课程"

add(B, "高等数学(1)", "4", "必修", {"1.1":"H","2.1":"H","2.2":"M"})
add(B, "高等数学(2)", "5", "必修", {"1.1":"H","2.1":"H","2.2":"M"})
add(B, "线性代数", "2.5", "必修", {"1.1":"H","2.2":"H"})
add(B, "概率论与数理统计", "3", "必修", {"1.1":"H","2.2":"H","4.2":"M"})
add(B, "复变函数与积分变换", "3", "必修", {"1.1":"H","2.2":"H"})
add(B, "大学物理(一)1", "3.5", "必修", {"1.2":"H","2.1":"M","4.1":"L"})
add(B, "大学物理(一)2", "3.5", "必修", {"1.2":"H","2.1":"M","4.1":"L"})
add(B, "大学物理实验(一)", "1", "必修", {"1.2":"M","3.2":"M","4.1":"H","4.2":"M","4.3":"M"})

B = "工程基础类课程"

add(B, "工程图学及电子线路CAD", "2", "必修", {"3.2":"M","5.1":"H","5.2":"M","10.1":"L"})
add(B, "计算机基础及高级语言程序设计", "4", "必修", {"3.3":"M","5.1":"H"})
add(B, "电路分析与电子线路(1)", "3.5", "必修", {"1.3":"H","2.3":"M","4.3":"H"})
add(B, "电路分析与电子线路(2)", "4", "必修", {"1.3":"H","2.3":"M","3.2":"H","4.3":"M"})
add(B, "数字逻辑与电路", "2.5", "必修", {"1.3":"M","2.3":"M","3.2":"M","5.2":"M"})
add(B, "信号与系统", "4", "必修", {"1.3":"H","1.4":"H","2.1":"H","2.2":"H","2.3":"H","4.2":"M"})
add(B, "电磁场与传输理论", "2.5", "必修", {"1.2":"H","1.3":"M","2.2":"H","3.2":"L","4.1":"M"})
add(B, "面向对象程序设计(限选)", "2.5", "限选", {"3.2":"M","3.3":"M","5.1":"H"})
add(B, "算法与数据结构(限选)", "2.5", "限选", {"2.3":"M","3.3":"M","5.1":"H"})

B = "专业基础类课程"

add(B, "通信工程专业导论与技术前沿", "0.5", "必修", {"1.4":"M","6.1":"H","12.2":"M"})
add(B, "计算机网络", "3", "必修", {"1.4":"L","2.1":"H","3.1":"M","5.1":"M"})
add(B, "微机原理与接口技术", "3.5", "必修", {"1.3":"M","3.3":"M","5.2":"H"})
add(B, "信息论与编码", "3", "必修", {"1.4":"H","2.2":"M","2.3":"H","3.1":"H"})
add(B, "现代通信原理", "4", "必修", {"1.4":"H","2.1":"H","3.2":"M","4.1":"M"})
add(B, "数字信号处理", "3", "必修", {"1.4":"H","2.3":"H","3.2":"M","4.3":"M"})
add(B, "移动通信", "2.5", "必修", {"1.4":"H","3.2":"M","4.1":"H","4.3":"M","6.1":"M"})
add(B, "科研方法与实践（双语）(限选)", "0.5", "限选", {"2.4":"H","4.1":"M","10.3":"M"})

B = "专业类课程"

add(B, "机器学习与深度学习基础(任选)", "3", "任选", {"1.4":"H","2.3":"M","3.3":"M","5.1":"H"})
add(B, "模式识别与机器视觉(任选)", "3", "任选", {"1.4":"M","3.2":"H","4.3":"M","5.1":"M"})
add(B, "光纤通信(任选)", "3", "任选", {"1.4":"H","3.2":"M","4.1":"H","4.3":"M"})
add(B, "无人机组网与低空通信(任选)", "3", "任选", {"1.4":"H","3.1":"M","3.4":"H","6.2":"M"})
add(B, "卫星通信与导航(任选)", "3", "任选", {"1.4":"H","3.2":"M","5.3":"M"})
add(B, "量子通信(任选)", "3", "任选", {"1.2":"H","1.4":"H","3.4":"M"})
add(B, "智能通信系统设计(任选)", "2", "任选", {"3.1":"H","3.3":"M","5.1":"M"})
add(B, "通信感知一体化设计(任选)", "2", "任选", {"3.4":"H","4.1":"M","5.3":"M"})

B = "工程实践与毕业设计"

add(B, "金工实习", "1", "必修", {"3.3":"H","5.2":"M","7.1":"L","9.1":"M","9.2":"M"})
add(B, "电装实习", "0.5", "必修", {"3.3":"H","5.2":"M"})
add(B, "电路分析与电子线路实验", "1", "必修", {"3.2":"H","4.1":"H","4.2":"M","5.2":"M","9.2":"M"})
add(B, "数字电路系统设计综合实践", "1", "必修", {"3.2":"H","3.3":"M","5.2":"H","9.2":"M"})
add(B, "电子技术课程设计", "1", "必修", {"3.1":"M","3.2":"H","3.3":"H","4.3":"L","5.2":"L","9.2":"M","10.1":"L","11.3":"L"})
add(B, "数字信号处理综合实践", "0.5", "必修", {"3.3":"H","4.2":"H","4.3":"M","5.1":"H"})
add(B, "移动通信仿真实践", "0.5", "必修", {"3.3":"M","4.2":"M","5.3":"H"})
add(B, "通信系统工程实践", "2", "必修", {"1.4":"M","3.2":"H","3.4":"M","6.2":"M","7.1":"H","9.2":"M","9.3":"M","11.3":"H"})
add(B, "专业实践训练", "2", "必修", {"4.1":"H","4.3":"M","5.1":"M","5.2":"M","9.1":"H","9.2":"M"})
add(B, "毕业实习", "4", "必修", {
    "4.1":"H","5.1":"H","5.2":"H","6.1":"M","6.2":"H","7.1":"M","7.2":"M",
    "8.1":"H","8.2":"M","9.1":"H","9.2":"H","9.3":"H",
    "10.1":"M","10.2":"M","11.1":"M","11.2":"M","11.3":"H"
})
add(B, "毕业设计（论文）", "6", "必修", {
    "2.4":"H","3.1":"H","3.2":"H","3.3":"H","3.4":"H",
    "4.1":"H","4.2":"H","4.3":"H","4.4":"H",
    "5.1":"H","5.3":"M","6.2":"M","7.2":"H",
    "8.3":"H","9.1":"H","9.3":"M",
    "10.1":"H","10.2":"M","10.3":"M",
    "11.1":"M","11.3":"M","12.1":"H","12.2":"H"
})

add(B, "Python程序设计综合实践(限选)", "1.5", "限选", {"3.3":"H","4.2":"M","5.1":"H"})
add(B, "软件系统设计综合实践(限选)", "0.5", "限选", {"3.3":"H","5.1":"M","9.2":"M"})
add(B, "通信系统软硬件协同设计实践(限选)", "0.5", "限选", {"3.3":"H","5.1":"M","5.2":"M"})
add(B, "FPGA应用开发(任选)", "3", "任选", {"3.2":"H","3.3":"H","5.2":"H"})
add(B, "机器人设计与制造(任选)", "3", "任选", {"3.2":"H","3.3":"H","5.2":"H","9.2":"M"})
add(B, "嵌入式系统开发(任选)", "3", "任选", {"3.3":"H","3.4":"M","5.2":"H"})
add(B, "机器人操作系统与编程(任选)", "3", "任选", {"3.3":"H","5.1":"H"})
add(B, "软件无线电技术应用(任选)", "3", "任选", {"3.2":"M","3.3":"H","4.1":"H","5.3":"H"})
add(B, "无人系统自主导航与规划实践(任选)", "3", "任选", {"3.2":"H","3.4":"H","5.1":"M"})
add(B, "人工智能教育技术(任选)", "2", "任选", {"3.3":"M","5.1":"M","12.2":"M"})
add(B, "通信网络安全与对抗实践(任选)", "1", "任选", {"3.3":"H","5.1":"H","6.2":"H","8.2":"M"})
add(B, "5G网络工程实践仿真(任选)", "1", "任选", {"3.3":"M","5.3":"H"})
add(B, "AI智能体开发(任选)", "1", "任选", {"3.3":"H","5.1":"H","12.2":"M"})

B = "自主发展课程"

add(B, "现代交换与网络(卓越工程师)", "3", "自选", {"1.4":"M","2.1":"H","3.2":"M","5.1":"H"})
add(B, "通信电子线路(卓越工程师)", "3", "自选", {"1.3":"M","1.4":"M","3.2":"M","4.1":"H","5.2":"M"})
add(B, "数字图像处理与分析(卓越工程师)", "2", "自选", {"1.4":"M","3.2":"H","4.2":"M","5.1":"M"})
add(B, "天线理论与设计(卓越工程师)", "2", "自选", {"1.4":"M","3.2":"H","4.1":"M","5.3":"H"})
add(B, "电路与电子学(学术深造)", "3", "自选", {"1.3":"H","2.3":"M","4.1":"M"})
add(B, "数学分析(学术深造)", "4", "自选", {"1.1":"H","2.2":"H","2.3":"H"})
add(B, "信号分析与处理(学术深造)", "3", "自选", {"1.4":"M","2.3":"H","4.2":"H","4.3":"M"})
add(B, "微专业课程群", "10", "自选", {"3.4":"M","12.2":"H"})

# ============================================================
# BUILD EXCEL
# ============================================================

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(name='微软雅黑', size=9, bold=True)
normal_font = Font(name='微软雅黑', size=8)
small_font = Font(name='微软雅黑', size=7)
title_font = Font(name='微软雅黑', size=14, bold=True)
H_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
M_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
L_fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Title
ws.merge_cells('A1:AR1')
ws['A1'] = '通信工程专业2026版课程体系与毕业要求支撑关系矩阵'
ws['A1'].font = title_font
ws['A1'].alignment = center_align

# Column layout
req_start_col = 5
col = req_start_col
for req_name, indicators in requirements.items():
    n = len(indicators)
    if n > 1:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+n-1)
    ws.cell(row=2, column=col, value=req_name).font = header_font
    ws.cell(row=2, column=col).fill = header_fill
    ws.cell(row=2, column=col).alignment = center_align
    ws.cell(row=2, column=col).border = thin_border
    for c in range(col, col+n):
        ws.cell(row=2, column=c).border = thin_border
    col += n

total_indicator_cols = col - req_start_col
last_col = req_start_col + total_indicator_cols - 1

# Fixed header columns
for cell_ref, label in [('A2', '序号'), ('B2', '课程板块'), ('C2', '课程名称'), ('D2', '学分')]:
    merge_start = cell_ref + ':' + cell_ref[0] + '4'
    ws.merge_cells(merge_start)
    ws[cell_ref] = label
    ws[cell_ref].font = header_font
    ws[cell_ref].fill = header_fill
    ws[cell_ref].alignment = center_align
    ws[cell_ref].border = thin_border

# Indicator codes (row 3) and descriptions (row 4)
col = req_start_col
for req_name, indicators in requirements.items():
    for code, desc in indicators:
        c = ws.cell(row=3, column=col, value=code)
        c.font = small_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

        c = ws.cell(row=4, column=col, value=desc)
        c.font = small_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
        col += 1

# Border all header cells
for r in range(2, 5):
    for c in range(1, last_col+1):
        ws.cell(row=r, column=c).border = thin_border

# Course data
row = 5
seq = 0
for (board, course_name, credit, nature, mapping) in courses:
    seq += 1
    c = ws.cell(row=row, column=1, value=seq)
    c.font = normal_font; c.alignment = center_align; c.border = thin_border

    c = ws.cell(row=row, column=2, value=board)
    c.font = normal_font; c.alignment = left_align; c.border = thin_border

    c = ws.cell(row=row, column=3, value=course_name)
    c.font = normal_font; c.alignment = left_align; c.border = thin_border

    c = ws.cell(row=row, column=4, value=credit)
    c.font = normal_font; c.alignment = center_align; c.border = thin_border

    col = req_start_col
    for req_name, indicators in requirements.items():
        for code, desc in indicators:
            if code in mapping:
                level = mapping[code]
                c = ws.cell(row=row, column=col, value=level)
                c.font = Font(name='微软雅黑', size=8, bold=True)
                c.alignment = center_align
                if level == 'H':
                    c.fill = H_fill
                elif level == 'M':
                    c.fill = M_fill
                elif level == 'L':
                    c.fill = L_fill
            ws.cell(row=row, column=col).border = thin_border
            col += 1
    row += 1

# Legend
row += 2
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws.cell(row=row, column=1, value='图例：').font = Font(name='微软雅黑', size=10, bold=True)
row += 1
for level, color_fill, label in [
    ('H', H_fill, '高支撑 — 课程直接支撑该指标点的达成，教学环节中明确覆盖并作为重点考核内容'),
    ('M', M_fill, '中支撑 — 课程部分支撑该指标点的达成，教学环节中有涉及并纳入考核'),
    ('L', L_fill, '低支撑 — 课程对该指标点有间接贡献，教学中有涉及但不作为重点考核')
]:
    ws.cell(row=row, column=1, value=level).font = Font(name='微软雅黑', size=9, bold=True)
    ws.cell(row=row, column=1).fill = color_fill
    ws.cell(row=row, column=1).alignment = center_align
    ws.cell(row=row, column=1).border = thin_border
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2, value=label).font = Font(name='微软雅黑', size=9)
    ws.cell(row=row, column=2).alignment = left_align
    row += 1

# Summary statistics
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws.cell(row=row, column=1, value='各毕业要求支撑课程数统计：').font = Font(name='微软雅黑', size=10, bold=True)
row += 1

# Summary header
for ci, label in [(1,'毕业要求'), (2,'H(门)'), (3,'M(门)'), (4,'L(门)'), (5,'合计')]:
    c = ws.cell(row=row, column=ci, value=label)
    c.font = header_font; c.fill = header_fill; c.border = thin_border; c.alignment = center_align
row += 1

for req_name, indicators in requirements.items():
    h_set = set(); m_set = set(); l_set = set()
    for _, cname, _, _, mapping in courses:
        for code, desc in indicators:
            if code in mapping:
                if mapping[code] == 'H': h_set.add(cname)
                elif mapping[code] == 'M': m_set.add(cname)
                elif mapping[code] == 'L': l_set.add(cname)
    hc, mc, lc = len(h_set), len(m_set), len(l_set)
    allc = len(h_set | m_set | l_set)
    for ci, val in [(1,req_name), (2,hc), (3,mc), (4,lc), (5,allc)]:
        c = ws.cell(row=row, column=ci, value=val)
        c.font = normal_font; c.border = thin_border
        if ci > 1:
            c.alignment = center_align
    row += 1

# Column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 36
ws.column_dimensions['D'].width = 5.5
for c in range(req_start_col, last_col+1):
    ws.column_dimensions[get_column_letter(c)].width = 6.5

# Row heights
ws.row_dimensions[1].height = 32
ws.row_dimensions[2].height = 28
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 42
for r in range(5, row):
    ws.row_dimensions[r].height = 22

# Freeze panes
ws.freeze_panes = 'E5'

# Print settings
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

output_path = "/Users/seansun/work/assistant folder/2026版通信工程专业课程体系与毕业要求支撑关系矩阵.xlsx"
wb.save(output_path)
print(f"✅ 已保存至: {output_path}")
print(f"课程总数: {len(courses)}")

# Verify coverage
print("\n=== 各指标点覆盖验证 ===")
indicator_coverage = {code: [] for code in all_indicators}
for _, cname, _, _, mapping in courses:
    for code, level in mapping.items():
        indicator_coverage[code].append((cname, level))

all_ok = True
for code in all_indicators:
    h = [c for c,l in indicator_coverage[code] if l=='H']
    m = [c for c,l in indicator_coverage[code] if l=='M']
    l = [c for c,l in indicator_coverage[code] if l=='L']
    status = "✅" if (h or m) else "⚠️"
    if status == "⚠️":
        all_ok = False
    line = f"  {code}: {status} H={len(h)}, M={len(m)}, L={len(l)}"
    if status == "⚠️":
        line += "  <-- 仅有L或无支撑!"
    print(line)

if all_ok:
    print("\n✅ 全部38个指标点均至少有1门H或M级课程支撑，满足工程教育认证要求")
else:
    print("\n⚠️ 部分指标点支撑不足，需调整")

print("\n=== 各毕业要求支撑统计 ===")
for req_name, indicators in requirements.items():
    h_set = set(); m_set = set(); l_set = set()
    for _, cname, _, _, mapping in courses:
        for code, desc in indicators:
            if code in mapping:
                if mapping[code] == 'H': h_set.add(cname)
                elif mapping[code] == 'M': m_set.add(cname)
                elif mapping[code] == 'L': l_set.add(cname)
    print(f"  {req_name}: H={len(h_set)}门, M={len(m_set)}门, L={len(l_set)}门, 合计={len(h_set|m_set|l_set)}门")
